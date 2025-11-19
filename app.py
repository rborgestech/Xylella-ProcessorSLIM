# -*- coding: utf-8 -*-
import io
import zipfile
from datetime import datetime
from typing import List, Tuple, Dict

import streamlit as st
from openpyxl import load_workbook

from processor import process_pre_to_dgav, REQUIRED_DGAV_COLS, _norm


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# AnÃ¡lise do ficheiro DGAV gerado
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def analyse_output_xlsx(xlsx_bytes: bytes) -> Tuple[int, List[str], List[str]]:
    """
    Analisa o ficheiro DGAV gerado:
      - sample_count: nÂº de amostras
      - warnings: avisos leves (nÃ£o usamos por agora)
      - hard_warnings: colunas obrigatÃ³rias com cÃ©lulas vazias/ausentes
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["Default"]

    header_indices: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            header_indices[_norm(v)] = col

    warnings: List[str] = []
    hard_warnings: List[str] = []

    # Determinar Ãºltima linha com dados (CODIGO_AMOSTRA)
    codigo_idx = header_indices.get(_norm("DESCRICAO"))
    sample_count = 0
    last_row = 1

    if codigo_idx:
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=codigo_idx).value
            if v not in (None, ""):
                sample_count += 1
                last_row = row
    else:
        hard_warnings.append("Coluna obrigatÃ³ria ausente no output: DESCRICAO")

    # Verificar colunas obrigatÃ³rias (modo 2: qualquer cÃ©lula vazia)
    for col_name in REQUIRED_DGAV_COLS:
        col_idx = header_indices.get(_norm(col_name))
        if col_idx is None:
            hard_warnings.append(f"Coluna obrigatÃ³ria ausente no output: {col_name}")
            continue

        for row in range(2, last_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v in (None, ""):
                hard_warnings.append(f"Coluna obrigatÃ³ria com cÃ©lulas vazias: {col_name}")
                break

    return sample_count, warnings, hard_warnings


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# ZIP com resultados e summary
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def build_zip_with_summary(
    outputs: List[Tuple[str, bytes, int, List[str], List[str], str]],
    summary_lines: List[str],
    timestamp: str,
) -> bytes:
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as z:
        for original_name, data, _, _, _, _ in outputs:
            base = original_name.rsplit(".", 1)[0]
            out_name = f"{base}_DGAV_{timestamp}.xlsx"
            z.writestr(out_name, data)

        z.writestr("summary.txt", "\n".join(summary_lines))

    mem.seek(0)
    return mem.getvalue()


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# UI styling
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
st.set_page_config(page_title="Xylella â†’ DGAV", page_icon="ğŸ§ª", layout="centered")

st.title("ğŸ§ª Xylella â€“ Conversor para SLIM")
st.caption(
    "Carrega ficheiros **â€œAVALIAÃ‡ÃƒO PRÃ‰-PROCESSAMENTO â€“ Amostras Xylellaâ€** "
    "e gera automaticamente o ficheiro SLIM **DGAV_SAMPLE_REGISTRATION_FILE_XYLELLA.xlsx**."
)

st.markdown(
    """
<style>
.stButton > button[kind="primary"]{
  background:#CA4300!important;border:1px solid #CA4300!important;color:#fff!important;
  font-weight:600!important;border-radius:6px!important;
}
.stButton > button[kind="primary"]:hover{
  background:#A13700!important;
}

/* upload box */
[data-testid="stFileUploader"]>div:first-child{
  border:2px dashed #CA4300!important;
  border-radius:10px!important;
  padding:1rem!important;
}

/* status boxes */
.file-box{
  border-radius:8px;
  padding:.6rem 1rem;
  margin-bottom:.5rem;
  opacity:0;
  animation:fadeIn .4s ease forwards;
}
@keyframes fadeIn{from{opacity:0;}to{opacity:1;}}

.file-box.processing{background:#E8F1FB;border-left:4px solid #2B6CB0;}
.file-box.success{background:#e6f9ee;border-left:4px solid #1a7f37;}
.file-box.warning{background:#fff8e5;border-left:4px solid #e6a100;}
.file-box.error{background:#fdeaea;border-left:4px solid #cc0000;}

.file-title{font-size:.9rem;font-weight:600;color:#1A365D;}
.file-sub{font-size:.8rem;color:#2A4365;}

/* loading dots */
.dots::after{
  content:'...';
  animation:dots 1.5s steps(4,end) infinite;
}
@keyframes dots{
  0%,20%{color:transparent;}
  40%{color:#2A4365;}
  60%{color:#2A4365;}
  80%,100%{color:#2A4365;}
}

/* clean button */
.clean-btn{
  background:#fff!important;border:1px solid #ccc!important;
  font-weight:600!important;border-radius:8px!important;
  padding:.5rem 1.2rem!important;
}
.clean-btn:hover{border-color:#999!important;}
</style>
""",
    unsafe_allow_html=True,
)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Session state
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
if "stage" not in st.session_state:
    st.session_state.stage = "idle"   # idle / processing / done
if "uploads" not in st.session_state:
    st.session_state.uploads = None
if "results" not in st.session_state:
    st.session_state.results = None


def reset_app():
    st.session_state.stage = "idle"
    st.session_state.uploads = None
    st.session_state.results = None


def render_results():
    """Mostra resultados jÃ¡ processados (fase 'done')."""
    res = st.session_state.results
    if not res:
        return

    outputs = res["outputs"]

    # Caixas por ficheiro
    for status in res["file_statuses"]:
        st.markdown(
            f"""
            <div class='file-box {status["status_class"]}'>
              <div class='file-title'>ğŸ“„ {status["name"]}</div>
              <div class='file-sub'><b>{status["sample_count"]}</b> amostra(s) processadas.</div>
              {status["message_html"]}
            </div>
            """,
            unsafe_allow_html=True,
        )

    # Resumo final
    st.markdown(
        f"""
        <div style='text-align:center;margin-top:1.5rem;'>
          <h3>ğŸ Processamento concluÃ­do!</h3>
          <p>Foram processados <b>{len(outputs)}</b> ficheiro(s) vÃ¡lido(s),
          com um total de <b>{res["total_samples"]}</b> amostras.<br>
          Ficheiros com avisos: <b>{res["warning_files"]}</b>.<br>
          Ficheiros com erro: <b>{res["error_files"]}</b>.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    timestamp = res["timestamp"]

    # BotÃµes de download
    if len(outputs) == 1:
        name, data, _, _, _, _ = outputs[0]
        base = name.rsplit(".", 1)[0]
        out_name = f"{base}_DGAV_{timestamp}.xlsx"

        st.download_button(
            "â¬‡ï¸ Descarregar ficheiro DGAV",
            data=data,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        zip_bytes = build_zip_with_summary(outputs, res["summary_lines"], timestamp)
        zip_name = f"xylella_dgav_{timestamp}.zip"
        st.download_button(
            "ğŸ“¦ Descarregar resultados (ZIP)",
            data=zip_bytes,
            file_name=zip_name,
            mime="application/zip",
        )

    st.markdown("<br>", unsafe_allow_html=True)
    st.button("ğŸ” Novo processamento", on_click=reset_app, use_container_width=True)


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# MÃ¡quina de estados: idle / processing / done
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
if st.session_state.stage == "idle":

    uploads = st.file_uploader(
        "ğŸ“‚ Carrega um ou vÃ¡rios ficheiros de prÃ©-registo (XLSX)",
        type=["xlsx"],
        accept_multiple_files=True,
    )

    if uploads:
        if st.button("ğŸ“„ Processar ficheiros de Input", type="primary"):
            st.session_state.uploads = uploads
            st.session_state.results = None
            st.session_state.stage = "processing"
            st.rerun()
    else:
        st.info("ğŸ’¡ Carrega pelo menos um ficheiro.")

elif st.session_state.stage == "processing":

    # Se por algum motivo jÃ¡ houver resultados, saltamos para 'done'
    if st.session_state.results is not None:
        st.session_state.stage = "done"
        st.rerun()

    uploads = st.session_state.uploads
    total = len(uploads)

    st.info("â³ A processar ficheiros... aguarde atÃ© o processo terminar.")
    progress = st.progress(0.0)

    outputs: List[Tuple[str, bytes, int, List[str], List[str], str]] = []
    summary_lines: List[str] = []
    file_statuses: List[Dict] = []
    total_samples = 0
    warning_files = 0
    error_files = 0
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    for i, up in enumerate(uploads, start=1):
        placeholder = st.empty()
        placeholder.markdown(
            f"""
            <div class='file-box processing'>
              <div class='file-title'>ğŸ“„ {up.name}</div>
              <div class='file-sub'>Ficheiro {i} de {total} â€” a processar<span class="dots"></span></div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        try:
            data_in = io.BytesIO(up.getbuffer())
            output_bytes, log_msg = process_pre_to_dgav(data_in)

            sample_count, soft_warns, hard_warns = analyse_output_xlsx(output_bytes)
            total_samples += sample_count

            combined_warns = hard_warns + soft_warns

            if combined_warns:
                status_class = "warning"
                warning_files += 1
                bullets = "<br>".join(f"â€¢ {w}" for w in combined_warns)
                message_html = f"<div class='file-sub'>âš ï¸ Avisos:<br>{bullets}</div>"
            else:
                status_class = "success"
                message_html = ""

            outputs.append(
                (up.name, output_bytes, sample_count, soft_warns, hard_warns, status_class)
            )

            summary_line = f"{up.name}: {sample_count} amostra(s). {log_msg}"
            if combined_warns:
                summary_line += " âš  " + " | ".join(combined_warns)
            summary_lines.append(summary_line)

            file_statuses.append(
                {
                    "name": up.name,
                    "sample_count": sample_count,
                    "status_class": status_class,
                    "message_html": message_html,
                }
            )

            placeholder.markdown(
                f"""
                <div class='file-box {status_class}'>
                  <div class='file-title'>ğŸ“„ {up.name}</div>
                  <div class='file-sub'><b>{sample_count}</b> amostra(s) processadas.</div>
                  {message_html}
                </div>
                """,
                unsafe_allow_html=True,
            )

        except Exception as e:
            error_files += 1
            msg = f"{up.name}: erro ao processar ({e})"
            summary_lines.append(msg)

            file_statuses.append(
                {
                    "name": up.name,
                    "sample_count": 0,
                    "status_class": "error",
                    "message_html": f"<div class='file-sub'>âŒ Erro ao processar: {e}</div>",
                }
            )

            placeholder.markdown(
                f"""
                <div class='file-box error'>
                  <div class='file-title'>ğŸ“„ {up.name}</div>
                  <div class='file-sub'>âŒ Erro ao processar: {e}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        progress.progress(i / total)

    summary_lines.append("")
    summary_lines.append(f"Total de ficheiros vÃ¡lidos: {len(outputs)}")
    summary_lines.append(f"Total de amostras: {total_samples}")
    summary_lines.append(f"Ficheiros com avisos: {warning_files}")
    summary_lines.append(f"Ficheiros com erro: {error_files}")
    summary_lines.append(f"Executado em: {datetime.now():%d/%m/%Y %H:%M:%S}")

    st.session_state.results = {
        "outputs": outputs,
        "summary_lines": summary_lines,
        "total_samples": total_samples,
        "warning_files": warning_files,
        "error_files": error_files,
        "timestamp": timestamp,
        "file_statuses": file_statuses,
    }

    # TransiÃ§Ã£o para fase 'done'
    st.session_state.stage = "done"
    st.rerun()

elif st.session_state.stage == "done":
    render_results()
