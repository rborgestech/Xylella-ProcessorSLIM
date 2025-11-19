# -*- coding: utf-8 -*-
from io import BytesIO
from pathlib import Path
from typing import Tuple, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import unicodedata


DGAV_TEMPLATE_PATH = Path(__file__).parent / "DGAV_SAMPLE_REGISTRATION_FILE_XYLELLA.xlsx"


INPUT_TO_DGAV_COLMAP = {
    "DATA_RECEPCAO": "Data recepção amostras",
    "DATA_COLHEITA": "Data colheita",
    "CODIGO_AMOSTRA": "Código_amostra (Código original / Referência amostra)",
    "HOSPEDEIRO": "Espécie indicada / Hospedeiro",
    "TIPO_AMOSTRA": "Tipo amostra Simples / Composta",
    "ID_ZONA": "Id Zona (Classificação de zona de origem)",
    "COD_INT_LAB": "Código interno Lab",
    "DATA_REQUERIDO": "Data requerido",
    "RESPONSAVEL_AMOSTRAGEM": "Responsável Amostragem (Zona colheita)",
    "RESP_COLHEITA": "Responsável colheita (Técnico responsável)",
    "PREP_COMMENTS": "Prep_Comments (Observações cliente)",
    "PROCEDURE": "Procedure",
}

REQUIRED_DGAV_COLS = [
    "DATA_RECEPCAO",
    "DATA_COLHEITA",
    "CODIGO_AMOSTRA",
    "HOSPEDEIRO",
    "TIPO_AMOSTRA",
    "ID_ZONA",
    "PROCEDURE",
    "COD_INT_LAB",
    "DATA_REQUERIDO",
]


# ───────────────────────────────────────────────
# Normalização "tolerante"
# ───────────────────────────────────────────────
def _norm(text: str | None) -> str:
    if text is None:
        return ""
    s = str(text)
    s = s.replace("\u00A0", " ")
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.replace("_", " ").replace("-", " ")
    s = s.lower()
    return " ".join(s.split()).strip()


# ───────────────────────────────────────────────
# Cabeçalho no pré-registo
# ───────────────────────────────────────────────
def _find_header_row(df_raw: pd.DataFrame, target: str) -> int:
    target_norm = _norm(target)

    # Passagem 1: match exato
    for idx, row in df_raw.iterrows():
        if row.astype(str).apply(_norm).eq(target_norm).any():
            return idx

    # Passagem 2: fallback "codigo amostra"
    for idx, row in df_raw.iterrows():
        if row.astype(str).apply(_norm).str.contains("codigo amostra", na=False).any():
            return idx

    raise ValueError("Não foi possível identificar a linha de cabeçalho no pré-registo.")


# ───────────────────────────────────────────────
# Ler pré-registo com fórmulas calculadas
# ───────────────────────────────────────────────
def _load_pre_registo_df(uploaded_file) -> pd.DataFrame:
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    rows = list(ws.values)
    df_raw = pd.DataFrame(rows)

    header_row = _find_header_row(
        df_raw, "Código_amostra (Código original / Referência amostra)"
    )
    headers = df_raw.iloc[header_row].tolist()

    df = df_raw.iloc[header_row + 1 :].copy()
    df.columns = headers
    df = df.dropna(how="all")

    return df


# ───────────────────────────────────────────────
# Filtrar apenas linhas com CODIGO_AMOSTRA
# ───────────────────────────────────────────────
def _filter_sample_rows(df: pd.DataFrame) -> pd.DataFrame:
    target_norm = _norm(INPUT_TO_DGAV_COLMAP["CODIGO_AMOSTRA"])
    cod_col = None

    for col in df.columns:
        if _norm(col) == target_norm:
            cod_col = col
            break

    if cod_col is None:
        return df  # fallback seguro

    mask = df[cod_col].notna() & (df[cod_col].astype(str).str.strip() != "")
    return df[mask].copy()


# ───────────────────────────────────────────────
# Mapear colunas pré-registo → DGAV
# ───────────────────────────────────────────────
def _map_input_columns(df: pd.DataFrame) -> Dict[str, str]:
    norm_to_real: Dict[str, str] = { _norm(col): col for col in df.columns }

    mapped: Dict[str, str] = {}
    for dgav_col, input_label in INPUT_TO_DGAV_COLMAP.items():
        key_norm = _norm(input_label)
        mapped[dgav_col] = norm_to_real.get(key_norm)

    return mapped


# ───────────────────────────────────────────────
# Cabeçalhos no template DGAV
# ───────────────────────────────────────────────
def _build_header_index(ws) -> Dict[str, int]:
    header_indices: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            header_indices[_norm(v)] = col
    return header_indices


# ───────────────────────────────────────────────
# Validar colunas obrigatórias
# ───────────────────────────────────────────────
def _mark_required_empty_columns(ws, header_indices: Dict[str, int], start_row: int, last_row: int):
    red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    for col_name in REQUIRED_DGAV_COLS:
        col_idx = header_indices.get(_norm(col_name))
        if col_idx is None:
            continue

        any_empty = False
        for r in range(start_row, last_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v in (None, ""):
                any_empty = True
                break

        if any_empty:
            ws.cell(row=1, column=col_idx).fill = red


# ───────────────────────────────────────────────
# PROCESSAMENTO PRINCIPAL
# ───────────────────────────────────────────────
def process_pre_to_dgav(uploaded_file) -> Tuple[bytes, str]:
    """
    Converte ficheiro de PRÉ-REGISTO → DGAV.
    Mantém listas de valores do template.
    Controla nº de linhas pelo CODIGO_AMOSTRA.
    """
    if not DGAV_TEMPLATE_PATH.exists():
        raise FileNotFoundError("Template DGAV não encontrado.")

    # 1) Carregar pré-registo
    df_in = _load_pre_registo_df(uploaded_file)

    # 🔹 IMPORTANTE: só linhas com código de amostra
    df_in = _filter_sample_rows(df_in)
    df_in = df_in.reset_index(drop=True)
    n_samples = len(df_in)

    input_colmap = _map_input_columns(df_in)

    # 2) Carregar template DGAV
    template_bytes = DGAV_TEMPLATE_PATH.read_bytes()
    template_stream = BytesIO(template_bytes)
    wb = load_workbook(template_stream)
    ws = wb["Default"]

    header_indices = _build_header_index(ws)

    # Guardar valores da linha 2 (defaults)
    base_values: Dict[str, object] = {}
    for norm_name, col_idx in header_indices.items():
        base_values[norm_name] = ws.cell(row=2, column=col_idx).value

    # 3) Apagar TODAS as linhas a partir da linha 2
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # 4) Escrever uma linha por amostra
    start_row = 2
    last_row = start_row + n_samples - 1 if n_samples > 0 else 1

    for i, (_, row_in) in enumerate(df_in.iterrows()):
        excel_row = start_row + i

        # 4.1 Preencher linha com defaults
        for norm_name, col_idx in header_indices.items():
            ws.cell(row=excel_row, column=col_idx).value = base_values.get(norm_name)

        # 4.2 Substituir colunas vindas do pré-registo
        for dgav_col, input_label in INPUT_TO_DGAV_COLMAP.items():
            col_idx = header_indices.get(_norm(dgav_col))
            if col_idx is None:
                continue

            df_col_name = input_colmap.get(dgav_col)
            if not df_col_name:
                # coluna não existe no pré-registo → mantém default
                continue

            value = row_in.get(df_col_name)

            # Converte NaN em None
            if isinstance(value, float) and pd.isna(value):
                value = None

            # Remove hora se for datetime
            if hasattr(value, "date"):
                value = value.date()

            ws.cell(row=excel_row, column=col_idx).value = value

    # 5) Validar colunas obrigatórias
    if n_samples > 0:
        _mark_required_empty_columns(ws, header_indices, start_row=start_row, last_row=last_row)

    # 6) (defensivo) Garantir que não há linhas extra
    if ws.max_row > last_row:
        ws.delete_rows(last_row + 1, ws.max_row - last_row)

    # 🔸 NÃO mexemos em ws.data_validations → listas do template mantêm-se

    # 7) Exportar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue(), f"Foram processadas {n_samples} amostras."
